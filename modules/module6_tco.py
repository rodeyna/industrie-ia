import requests
import openpyxl
from openpyxl.styles import Font, Alignment
import os

def calculate_tco_node(state: dict):
    """
    Module 6: Financial Brain
    1. Fetches real-time JSON data (World Bank).
    2. Calculates 10-year TCO.
    3. Exports detailed data to Excel.
    """
    print("\n--- [NODE] Executing Module 6: Economic Intelligence ---")
    
    # JSON data fetch (World Bank)
    inflation_rate = 9.3  # Standard fallback for Algerian context
    try:
        # Fetching Inflation, consumer prices (annual %) for Algeria
        wb_url = "https://api.worldbank.org/v2/country/DZ/indicator/FP.CPI.TOTL.ZG?format=json"
        response = requests.get(wb_url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            # Loop to find the most recent valid year
            for entry in data[1]:
                if entry['value'] is not None:
                    inflation_rate = entry['value']
                    break
    except Exception as e:
        print(f"⚠️ Could not fetch live economic data: {e}")

    # TCO CALCULATION LOGIC
    specs = state.get("specs", {})
    # Use a default base price if Module 1/2 didn't provide one
    base_price = 500000 
    annual_maint_base = base_price * 0.07  # 7% of purchase price
    
    projection_years = 10
    maintenance_schedule = []
    accumulated_cost = base_price

    for year in range(1, projection_years + 1):
        # Maintenance grows by inflation annually
        yearly_cost = annual_maint_base * ((1 + (inflation_rate / 100)) ** year)
        maintenance_schedule.append(round(yearly_cost, 2))
        accumulated_cost += yearly_cost

    # EXCEL EXPORT (OpenPyXL) 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Analysis"
    
    # Build Header
    headers = ["Year", "Cost Type", "Amount (DZD)", "Inflation Factor"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Add Data Rows
    ws.append([0, "CAPEX (Initial Purchase)", base_price, "N/A"])
    for i, cost in enumerate(maintenance_schedule, 1):
        ws.append([i, "OPEX (Maintenance)", cost, f"{inflation_rate:.2f}%"])

    # Final Total Row
    ws.append([])
    ws.append(["TOTAL TCO", "", round(accumulated_cost, 2)])
    ws[ws.max_row][0].font = Font(bold=True)

    excel_path = "outputs/tco_analysis.xlsx"
    os.makedirs("outputs", exist_ok=True)
    wb.save(excel_path)

    # RETURN STATE 
    tco_results = {
        "total": f"{accumulated_cost:,.2f}",
        "inflation_rate": f"{inflation_rate:.2f}%",
        "excel_path": excel_path,
        "currency": "DZD"
    }

    return {
    **state,                  # ← ADD THIS
    "tco_results": tco_results,
    "tco": tco_results,       # ← ADD THIS (so M8/M9 can read it)
    "logs": [f"TCO calculated"]
}
run_module6 = calculate_tco_node